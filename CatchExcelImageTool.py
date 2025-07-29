#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CatchExcelImage.py
支持从Excel中提取两种类型的图片：
1. WPS的DISPIMG嵌入图片（通过cellimages.xml）
2. Office的浮动图片（通过xl/media目录和drawings.xml）

支持四种提取粒度：
    1. 整个工作簿
    2. 指定工作表
    3. 指定工作表某一列
    4. 通过图片ID直接提取
"""

from __future__ import annotations
import os
import zipfile
import xml.etree.ElementTree as ET
import openpyxl
from typing import Dict, List, Tuple, Optional
import re

# 全局缓存：存储工作表的图片位置映射
_image_position_cache = {}
_workbook_cache = {}

# ----------------------------------------------------------
# 内部工具函数
# ----------------------------------------------------------

def _build_image_position_cache(xlsx_path: str, sheet_name: str) -> Dict[Tuple[int, int], int]:
    """
    构建图片位置缓存，返回 {(row, col): image_index} 的映射
    只在第一次访问时构建，后续直接使用缓存
    """
    cache_key = f"{xlsx_path}#{sheet_name}"
    
    if cache_key in _image_position_cache:
        return _image_position_cache[cache_key]
    
    position_map = {}
    
    try:
        # 检查是否已经有缓存的workbook
        if xlsx_path in _workbook_cache:
            wb = _workbook_cache[xlsx_path]
        else:
            wb = openpyxl.load_workbook(xlsx_path, data_only=False)
            _workbook_cache[xlsx_path] = wb
        
        if sheet_name not in wb.sheetnames:
            _image_position_cache[cache_key] = position_map
            return position_map
        
        ws = wb[sheet_name]
        
        if hasattr(ws, '_images') and ws._images:
            print(f"缓存构建: 发现 {len(ws._images)} 个图片，正在构建位置映射...")
            
            for i, img in enumerate(ws._images):
                try:
                    if hasattr(img, 'anchor'):
                        anchor = img.anchor
                        
                        img_row = None
                        img_col = None
                        
                        # 检查不同类型的锚点
                        if hasattr(anchor, '_from'):
                            anchor_from = anchor._from
                            if hasattr(anchor_from, 'row') and hasattr(anchor_from, 'col'):
                                img_row = anchor_from.row + 1  # 转换为1基索引
                                img_col = anchor_from.col + 1
                        elif hasattr(anchor, 'col') and hasattr(anchor, 'row'):
                            img_row = anchor.row + 1
                            img_col = anchor.col + 1
                        
                        if img_row is not None and img_col is not None:
                            position_map[(img_row, img_col)] = i
                            print(f"缓存构建: 图片{i} -> 位置({img_row}, {img_col})")
                        
                except Exception as e:
                    print(f"缓存构建: 处理图片{i}时出错: {e}")
                    continue
            
            print(f"缓存构建完成: 共映射 {len(position_map)} 个图片位置")
        
    except Exception as e:
        print(f"构建图片位置缓存时出错: {e}")
    
    _image_position_cache[cache_key] = position_map
    return position_map

def _get_cached_workbook(xlsx_path: str):
    """
    获取缓存的workbook，避免重复加载
    """
    if xlsx_path not in _workbook_cache:
        _workbook_cache[xlsx_path] = openpyxl.load_workbook(xlsx_path, data_only=False)
    return _workbook_cache[xlsx_path]

def clear_image_cache():
    """
    清除图片缓存（用于释放内存或重新加载）
    """
    global _image_position_cache, _workbook_cache
    
    # 关闭所有缓存的workbook
    for wb in _workbook_cache.values():
        try:
            wb.close()
        except:
            pass
    
    _image_position_cache.clear()
    _workbook_cache.clear()
    print("图片缓存已清除")
def _extract_dispimg_ids(ws: openpyxl.worksheet.worksheet.Worksheet,
                         target_col: Optional[str] = None) -> List[str]:
    """
    从 openpyxl 工作表对象里提取所有 DISPIMG 的图片 ID。
    如果指定了 target_col（如 'A'），则只扫描该列。
    """
    ids = []
    col_idx = openpyxl.utils.column_index_from_string(target_col) if target_col else None

    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if target_col and cell.column != col_idx:
                continue
            value = str(cell.value or "")
            if '=_xlfn.DISPIMG(' in value:
                start = value.find('"') + 1
                end = value.find('"', start)
                ids.append(value[start:end])
    return ids


def _build_id_to_image_map(xlsx_path: str) -> Dict[str, str]:
    """
    解析 .xlsx 内部结构，返回 {image_id -> image内部路径} 的映射（DISPIMG图片）
    """
    try:
        with zipfile.ZipFile(xlsx_path, 'r') as z:
            cellimages_xml = z.read('xl/cellimages.xml')
            rels_xml = z.read('xl/_rels/cellimages.xml.rels')
    except KeyError:
        # 如果没有cellimages.xml，说明没有DISPIMG图片
        return {}

    root = ET.fromstring(cellimages_xml)
    root_rels = ET.fromstring(rels_xml)

    namespaces = {
        'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
        'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
        'r': 'http://schemas.openxmlformats.org/package/2006/relationships'
    }

    # 1. name -> rid
    name_to_rid = {}
    for pic in root.findall('.//xdr:pic', namespaces):
        name = pic.find('.//xdr:cNvPr', namespaces).attrib['name']
        rid = pic.find('.//a:blip', namespaces).attrib[
            '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed']
        name_to_rid[name] = rid

    # 2. rid -> 内部路径
    rid_to_path = {}
    for rel in root_rels.findall('.//r:Relationship', namespaces):
        rid_to_path[rel.attrib['Id']] = rel.attrib['Target']

    return {name: rid_to_path[rid] for name, rid in name_to_rid.items() if rid in rid_to_path}


def _extract_floating_images(xlsx_path: str) -> List[str]:
    """
    提取Excel中的浮动图片（Office格式），返回图片文件名列表
    """
    floating_images = []
    
    try:
        with zipfile.ZipFile(xlsx_path, 'r') as z:
            # 检查xl/media目录中的图片文件
            for file_info in z.filelist:
                if file_info.filename.startswith('xl/media/') and not file_info.is_dir():
                    # 提取文件名（不包含路径）
                    image_filename = os.path.basename(file_info.filename)
                    floating_images.append(image_filename)
    except Exception:
        # 如果没有xl/media目录或其他错误，返回空列表
        pass
    
    return floating_images


def _get_cell_floating_image(xlsx_path: str, sheet_name: str, cell_address: str) -> Optional[str]:
    """
    获取指定单元格位置的浮动图片
    通过分析drawings.xml来确定图片与单元格的位置关系
    """
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=False)
        if sheet_name not in wb.sheetnames:
            return None
        
        ws = wb[sheet_name]
        
        # 获取单元格的行列信息
        cell = ws[cell_address]
        target_row = cell.row
        target_col = cell.column
        
        # 检查工作表是否有图片
        if hasattr(ws, '_images') and ws._images:
            # 遍历工作表中的图片
            for img in ws._images:
                if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                    # 获取图片锚点信息
                    anchor_from = img.anchor._from
                    if hasattr(anchor_from, 'row') and hasattr(anchor_from, 'col'):
                        # 检查图片是否在目标单元格附近（允许一定的偏差）
                        img_row = anchor_from.row + 1  # openpyxl使用0基索引
                        img_col = anchor_from.col + 1
                        
                        # 如果图片位置与目标单元格精确匹配（0偏差）
                        if img_row == target_row and img_col == target_col:
                            # 尝试从图片对象中获取文件名
                            if hasattr(img, '_id') or hasattr(img, 'path'):
                                # 这里需要进一步解析来获取实际的图片文件名
                                # 由于openpyxl的限制，我们使用备用方案
                                pass
        
        wb.close()
        
    except Exception:
        pass
    
    return None


# ----------------------------------------------------------
# 对外 API
# ----------------------------------------------------------
def extract_workbook_images(xlsx_path: str,
                            output_dir: str = 'images') -> List[str]:
    """
    提取整个工作簿里所有 DISPIMG 图片。
    返回已保存图片的绝对路径列表。
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    all_ids = []
    for ws in wb.worksheets:
        all_ids.extend(_extract_dispimg_ids(ws))

    id_to_path = _build_id_to_image_map(xlsx_path)
    os.makedirs(output_dir, exist_ok=True)

    saved = []
    with zipfile.ZipFile(xlsx_path, 'r') as z:
        for img_id in set(all_ids):
            if img_id in id_to_path:
                img_internal = 'xl/' + id_to_path[img_id]
                img_data = z.read(img_internal)
                out_file = os.path.join(output_dir, f"{img_id}.png")
                with open(out_file, 'wb') as f:
                    f.write(img_data)
                saved.append(os.path.abspath(out_file))
    return saved


def extract_sheet_images(xlsx_path: str,
                         sheet_name: str,
                         output_dir: str = 'images') -> List[str]:
    """
    仅提取指定工作表中的 DISPIMG 图片。
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    ws = wb[sheet_name]
    ids = _extract_dispimg_ids(ws)

    id_to_path = _build_id_to_image_map(xlsx_path)
    os.makedirs(output_dir, exist_ok=True)

    saved = []
    with zipfile.ZipFile(xlsx_path, 'r') as z:
        for img_id in set(ids):
            if img_id in id_to_path:
                img_internal = 'xl/' + id_to_path[img_id]
                img_data = z.read(img_internal)
                out_file = os.path.join(output_dir, f"{img_id}.png")
                with open(out_file, 'wb') as f:
                    f.write(img_data)
                saved.append(os.path.abspath(out_file))
    return saved


def extract_column_images(xlsx_path: str,
                          sheet_name: str,
                          column: str,
                          output_dir: str = 'images') -> List[str]:
    """
    仅提取指定工作表某一列中的 DISPIMG 图片。
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    ws = wb[sheet_name]
    ids = _extract_dispimg_ids(ws, target_col=column)

    id_to_path = _build_id_to_image_map(xlsx_path)
    os.makedirs(output_dir, exist_ok=True)

    saved = []
    with zipfile.ZipFile(xlsx_path, 'r') as z:
        for img_id in set(ids):
            if img_id in id_to_path:
                img_internal = 'xl/' + id_to_path[img_id]
                img_data = z.read(img_internal)
                out_file = os.path.join(output_dir, f"{img_id}.png")
                with open(out_file, 'wb') as f:
                    f.write(img_data)
                saved.append(os.path.abspath(out_file))
    return saved


def extract_floating_images_from_sheet(xlsx_path: str,
                                       sheet_name: str,
                                       output_dir: str = 'images') -> List[str]:
    """
    提取指定工作表中的所有浮动图片
    返回已保存图片的绝对路径列表
    """
    os.makedirs(output_dir, exist_ok=True)
    saved_images = []
    
    try:
        with zipfile.ZipFile(xlsx_path, 'r') as z:
            # 获取所有xl/media目录中的图片
            media_files = []
            for file_info in z.filelist:
                if file_info.filename.startswith('xl/media/') and not file_info.is_dir():
                    # 检查是否是图片文件
                    ext = os.path.splitext(file_info.filename)[1].lower()
                    if ext in ['.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff', '.webp']:
                        media_files.append(file_info.filename)
            
            # 提取每个图片文件
            for i, media_file in enumerate(media_files):
                try:
                    img_data = z.read(media_file)
                    # 获取文件扩展名
                    file_ext = os.path.splitext(media_file)[1] or '.png'
                    # 使用原始文件名作为输出文件名
                    original_name = os.path.basename(media_file)
                    out_file = os.path.join(output_dir, f"floating_{i+1}_{original_name}")
                    
                    with open(out_file, 'wb') as f:
                        f.write(img_data)
                    saved_images.append(os.path.abspath(out_file))
                    print(f"成功提取浮动图片: {out_file}")
                except Exception as e:
                    print(f"提取图片 {media_file} 时出错: {e}")
                    continue
                    
    except Exception as e:
        print(f"提取浮动图片时出错: {e}")
        pass
    
    return saved_images


def extract_first_available_image(xlsx_path: str,
                                 output_dir: str = 'images',
                                 prefix: str = 'image') -> Optional[str]:
    """
    提取第一个可用的图片（DISPIMG或浮动图片）
    这是一个简化的备用方案
    """
    os.makedirs(output_dir, exist_ok=True)
    
    # 1. 先尝试DISPIMG图片
    try:
        dispimg_ids = get_all_image_ids(xlsx_path)
        if dispimg_ids:
            return extract_image_by_id(xlsx_path, dispimg_ids[0], output_dir)
    except Exception:
        pass
    
    # 2. 再尝试浮动图片
    try:
        with zipfile.ZipFile(xlsx_path, 'r') as z:
            for file_info in z.filelist:
                if file_info.filename.startswith('xl/media/') and not file_info.is_dir():
                    ext = os.path.splitext(file_info.filename)[1].lower()
                    if ext in ['.png', '.jpg', '.jpeg', '.gif', '.bmp']:
                        img_data = z.read(file_info.filename)
                        out_file = os.path.join(output_dir, f"{prefix}{ext}")
                        
                        with open(out_file, 'wb') as f:
                            f.write(img_data)
                        return os.path.abspath(out_file)
    except Exception:
        pass
    
    return None


def extract_image_by_id(xlsx_path: str,
                        image_id: str,
                        output_dir: str = 'images') -> Optional[str]:
    """
    通过指定的图片ID直接提取图片（DISPIMG格式）。
    返回保存的图片绝对路径，如果图片ID不存在则返回None。
    """
    id_to_path = _build_id_to_image_map(xlsx_path)
    
    if image_id not in id_to_path:
        return None
    
    os.makedirs(output_dir, exist_ok=True)
    
    with zipfile.ZipFile(xlsx_path, 'r') as z:
        img_internal = 'xl/' + id_to_path[image_id]
        img_data = z.read(img_internal)
        out_file = os.path.join(output_dir, f"{image_id}.png")
        with open(out_file, 'wb') as f:
            f.write(img_data)
        return os.path.abspath(out_file)


def extract_image_from_cell(xlsx_path: str,
                           sheet_name: str,
                           cell_address: str,
                           output_dir: str = 'images') -> Optional[str]:
    """
    从指定单元格提取图片（优化版本，使用缓存）
    
    Args:
        xlsx_path: Excel文件路径
        sheet_name: 工作表名称
        cell_address: 单元格地址，如 'A1', 'B2'
        output_dir: 输出目录
    
    Returns:
        提取的图片文件路径，如果没有找到图片则返回None
    """
    try:
        # 使用缓存的工作簿
        wb = _get_cached_workbook(xlsx_path)
        
        if sheet_name not in wb.sheetnames:
            print(f"工作表 '{sheet_name}' 不存在")
            return None
        
        ws = wb[sheet_name]
        cell = ws[cell_address]
        
        # 首先尝试提取DISPIMG格式的图片
        # 从单元格地址中提取列字母（如从'N4'提取'N'）
        cell_column = openpyxl.utils.get_column_letter(cell.column)
        dispimg_ids = _extract_dispimg_ids(ws, cell_column)
        if dispimg_ids:
            print(f"在单元格 {cell_address} 找到DISPIMG图片: {dispimg_ids}")
            for img_id in dispimg_ids:
                result = extract_image_by_id(xlsx_path, img_id, output_dir)
                if result:
                    return result
        
        # 如果没有DISPIMG，尝试提取浮动图片（使用优化的缓存方法）
        print(f"在单元格 {cell_address} 未找到DISPIMG，尝试浮动图片...")
        result = _extract_floating_image_from_cell(xlsx_path, ws, cell.row, cell.column, output_dir)
        if result:
            return result
        
        print(f"在单元格 {cell_address} 未找到任何图片")
        return None
        
    except Exception as e:
        print(f"提取图片时出错: {e}")
        return None


def _extract_floating_image_from_cell(xlsx_path: str, ws, target_row: int, target_col: int, output_dir: str) -> Optional[str]:
    """
    从指定单元格位置提取浮动图片的内部方法（优化版本，使用缓存）
    """
    os.makedirs(output_dir, exist_ok=True)
    
    # 获取工作表名称
    sheet_name = ws.title
    
    # 使用缓存快速查找图片位置
    position_map = _build_image_position_cache(xlsx_path, sheet_name)
    
    print(f"快速查找: 目标位置({target_row}, {target_col})")
    
    # 直接查找精确匹配的图片
    if (target_row, target_col) in position_map:
        image_index = position_map[(target_row, target_col)]
        print(f"快速匹配: 在位置({target_row}, {target_col})找到图片{image_index}")
        
        # 获取图片对象并提取
        if hasattr(ws, '_images') and ws._images and image_index < len(ws._images):
            img = ws._images[image_index]
            return _extract_image_from_openpyxl_object(img, image_index, output_dir)
    
    print(f"快速查找: 位置({target_row}, {target_col})没有找到精确匹配的图片")
    return None
    
    # 如果无法精确定位图片位置，不提取任何图片
    # 这样可以避免提取不相关的图片
    
    return None


def _extract_image_from_openpyxl_object(img, image_index: int, output_dir: str) -> Optional[str]:
    """
    从openpyxl图片对象中提取图片数据
    """
    try:
        os.makedirs(output_dir, exist_ok=True)
        
        # 尝试获取图片数据
        if hasattr(img, '_data'):
            img_data = img._data()
        elif hasattr(img, 'ref'):
            # 有些版本的openpyxl使用ref属性
            img_data = img.ref
        else:
            print(f"调试: 无法从图片对象获取数据，图片属性: {dir(img)}")
            return None
        
        # 确定文件扩展名
        file_ext = '.png'  # 默认扩展名
        if hasattr(img, 'format') and img.format:
            file_ext = f'.{img.format.lower()}'
        
        out_file = os.path.join(output_dir, f"matched_image_{image_index+1}{file_ext}")
        
        with open(out_file, 'wb') as f:
            f.write(img_data)
        
        print(f"调试: 成功提取图片到 {out_file}")
        return os.path.abspath(out_file)
        
    except Exception as e:
        print(f"调试: 从openpyxl对象提取图片失败: {e}")
        return None


def _extract_specific_floating_image(xlsx_path: str, image_index: int, output_dir: str) -> Optional[str]:
    """
    提取指定索引的浮动图片
    """
    try:
        with zipfile.ZipFile(xlsx_path, 'r') as z:
            media_files = []
            for file_info in z.filelist:
                if file_info.filename.startswith('xl/media/') and not file_info.is_dir():
                    ext = os.path.splitext(file_info.filename)[1].lower()
                    if ext in ['.png', '.jpg', '.jpeg', '.gif', '.bmp']:
                        media_files.append(file_info.filename)
            
            if image_index < len(media_files):
                media_file = media_files[image_index]
                img_data = z.read(media_file)
                file_ext = os.path.splitext(media_file)[1] or '.png'
                out_file = os.path.join(output_dir, f"floating_image_{image_index+1}{file_ext}")
                
                with open(out_file, 'wb') as f:
                    f.write(img_data)
                return os.path.abspath(out_file)
            
    except Exception as e:
        print(f"提取特定浮动图片失败: {e}")
        pass
    
    return None


def get_all_image_ids(xlsx_path: str) -> List[str]:
    """
    获取Excel文件中所有可用的图片ID列表。
    返回所有图片ID的列表。
    """
    id_to_path = _build_id_to_image_map(xlsx_path)
    return list(id_to_path.keys())


# ----------------------------------------------------------
# 演示入口
# ----------------------------------------------------------
def main() -> None:
    """演示多种图片提取功能，支持DISPIMG和浮动图片"""
    xlsx = '（打印）VM销售策划.xlsx'

    # 1) 提取整个工作簿的DISPIMG图片
    print("【模式1】提取整个工作簿DISPIMG图片...")
    saved1 = extract_workbook_images(xlsx, './resources/images/output_all')
    print("已保存：", saved1)

    # 2) 提取指定工作表的DISPIMG图片
    print("\n【模式2】提取 'Sheet1' 工作表DISPIMG图片...")
    saved2 = extract_sheet_images(xlsx, 'Sheet1', './resources/images/output_sheet')
    print("已保存：", saved2)

    # 3) 提取 'Sheet2' 的 B 列DISPIMG图片
    print("\n【模式3】提取 'Sheet2' 的 B 列DISPIMG图片...")
    saved3 = extract_column_images(xlsx, 'Sheet2', 'B', './resources/images/output_column')
    print("已保存：", saved3)

    # 4) 提取浮动图片
    print("\n【模式4】提取 'Sheet1' 工作表的浮动图片...")
    saved4 = extract_floating_images_from_sheet(xlsx, 'Sheet1', './resources/images/output_floating')
    print("已保存：", saved4)

    # 5) 获取所有DISPIMG图片ID
    print("\n【辅助功能】获取所有DISPIMG图片ID...")
    all_ids = get_all_image_ids(xlsx)
    print("可用的图片ID：", all_ids)
    
    # 6) 通过ID直接提取DISPIMG图片
    print("\n【模式5】通过ID直接提取DISPIMG图片...")
    if all_ids:
        # 使用第一个可用的图片ID作为示例
        test_id = all_ids[0]
        print(f"使用图片ID: {test_id}")
        saved5 = extract_image_by_id(xlsx, test_id, './resources/images/output_by_id')
        if saved5:
            print("已保存：", saved5)
        else:
            print(f"图片ID '{test_id}' 提取失败")
    else:
        print("没有找到任何DISPIMG图片ID")
    
    # 7) 从指定单元格自动提取图片（DISPIMG或浮动）
    print("\n【模式6】从指定单元格自动提取图片...")
    saved6 = extract_image_from_cell(xlsx, 'Sheet1', 'B2', './resources/images/output_cell')
    if saved6:
        print("已保存：", saved6)
    else:
        print("单元格B2没有找到图片")
    
    print("\n=== 功能说明 ===")
    print("1. DISPIMG图片：WPS特有的嵌入图片格式，通过=DISPIMG()函数引用")
    print("2. 浮动图片：Office标准的图片插入方式，图片浮动在工作表上")
    print("3. 自动检测：extract_image_from_cell()函数会自动判断图片类型并提取")
    print("4. 兼容性：支持WPS和Office两种Excel格式的图片提取")


if __name__ == '__main__':
    main()